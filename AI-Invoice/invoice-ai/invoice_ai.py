from __future__ import annotations

import argparse
import json
import os
from collections import defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT / "invoice_ai_out"
OUT_DIR.mkdir(exist_ok=True)

TWO = Decimal("0.01")

# Valid work categories for line item classification
CATEGORIES = ["development", "consulting", "testing", "meeting", "support", "other"]


# ----------------------------
# Helpers
# ----------------------------
def to_date(x: Any) -> Optional[date]:
    if x is None:
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    s = str(x).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def iso(d: Optional[date]) -> Optional[str]:
    return d.isoformat() if d else None


def dec(x: Any) -> Optional[Decimal]:
    if x is None or x == "":
        return None
    try:
        return Decimal(str(x))
    except Exception:
        return None


def money_eur(x: Decimal) -> str:
    q = x.quantize(TWO, rounding=ROUND_HALF_UP)
    return f"€{q:.2f}".replace(".", ",")


def money_float(x: Decimal) -> float:
    return float(x.quantize(TWO, rounding=ROUND_HALF_UP))


def strip_code_fences(s: str) -> str:
    s = (s or "").strip()
    if s.startswith("```"):
        s = s.split("\n", 1)[1] if "\n" in s else ""
        if s.rstrip().endswith("```"):
            s = s.rsplit("```", 1)[0]
    return s.strip()


def call_llm_text(client: OpenAI, model: str, prompt: str) -> str:
    resp = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        temperature=0,
    )
    return (resp.choices[0].message.content or "").strip()


def call_llm_json(client: OpenAI, model: str, prompt: str) -> Dict[str, Any]:
    resp = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        temperature=0,
    )
    raw = (resp.choices[0].message.content or "").strip()
    raw = strip_code_fences(raw)
    return json.loads(raw)


# ----------------------------
# Data model
# ----------------------------
@dataclass
class InvoiceHeader:
    factuurnummer: str
    factuurdatum: Optional[date]
    vervaldatum: Optional[date]
    debiteurnummer: str
    klantnaam: str
    klant_adres: str
    klant_postcode_plaats: str


@dataclass
class InvoiceLine:
    factuurnummer: str
    omschrijving: str
    datum: Optional[date]
    aantal_uren: Optional[Decimal]
    tarief: Optional[Decimal]
    btw_pct: Optional[Decimal]


# ----------------------------
# Excel loading
# ----------------------------
def load_excel(path: Path) -> Tuple[List[InvoiceHeader], List[InvoiceLine]]:
    wb = load_workbook(path, data_only=True)
    inv_ws = wb["Facturen"]
    reg_ws = wb["Regels"]

    inv_header = [c.value for c in next(inv_ws.iter_rows(min_row=1, max_row=1))]
    reg_header = [c.value for c in next(reg_ws.iter_rows(min_row=1, max_row=1))]

    def row_to_dict(header: List[Any], row: Tuple[Any, ...]) -> Dict[str, Any]:
        d = {}
        for i, key in enumerate(header):
            if key is None:
                continue
            d[str(key).strip()] = row[i] if i < len(row) else None
        return d

    invoices: List[InvoiceHeader] = []
    for row in inv_ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        d = row_to_dict(inv_header, row)
        invoices.append(
            InvoiceHeader(
                factuurnummer=str(d.get("Factuurnummer") or "").strip(),
                factuurdatum=to_date(d.get("Factuurdatum")),
                vervaldatum=to_date(d.get("Vervaldatum")),
                debiteurnummer=str(d.get("Debiteurnummer") or "").strip(),
                klantnaam=str(d.get("Klantnaam") or "").strip(),
                klant_adres=str(d.get("Straat + huisnr") or "").strip(),
                klant_postcode_plaats=str(d.get("Postcode + Plaats") or "").strip(),
            )
        )

    lines: List[InvoiceLine] = []
    for row in reg_ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        d = row_to_dict(reg_header, row)
        lines.append(
            InvoiceLine(
                factuurnummer=str(d.get("Factuurnummer") or "").strip(),
                omschrijving=str(d.get("Omschrijving") or "").strip(),
                datum=to_date(d.get("Datum")),
                aantal_uren=dec(d.get("Aantal uren")),
                tarief=dec(d.get("Tarief")),
                btw_pct=dec(d.get("BTW_percentage")),
            )
        )

    return invoices, lines


def group_lines(lines: List[InvoiceLine]) -> Dict[str, List[InvoiceLine]]:
    out: Dict[str, List[InvoiceLine]] = {}
    for ln in lines:
        out.setdefault(ln.factuurnummer, []).append(ln)
    return out


def group_by_client(invoices: List[InvoiceHeader]) -> Dict[str, List[InvoiceHeader]]:
    out: Dict[str, List[InvoiceHeader]] = {}
    for inv in invoices:
        out.setdefault(inv.klantnaam, []).append(inv)
    return out


# ----------------------------
# Deterministic calculations
# ----------------------------
def compute_totals(lines: List[InvoiceLine]) -> Dict[str, Any]:
    net_total = Decimal("0.00")
    vat_total = Decimal("0.00")
    total_hours = Decimal("0.00")

    for ln in lines:
        if ln.aantal_uren is None or ln.tarief is None:
            continue
        net = (ln.aantal_uren * ln.tarief).quantize(TWO, rounding=ROUND_HALF_UP)
        net_total += net
        total_hours += ln.aantal_uren
        if ln.btw_pct is not None:
            vat = (net * (ln.btw_pct / Decimal("100"))).quantize(TWO, rounding=ROUND_HALF_UP)
            vat_total += vat

    gross = (net_total + vat_total).quantize(TWO, rounding=ROUND_HALF_UP)

    return {
        "totaal_excl": money_eur(net_total),
        "totaal_excl_float": money_float(net_total),
        "btw_totaal": money_eur(vat_total),
        "totaal_incl": money_eur(gross),
        "totaal_incl_float": money_float(gross),
        "totaal_uren": float(total_hours),
    }


def invoice_to_dict(inv: InvoiceHeader, lines: List[InvoiceLine], totals: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "factuurnummer": inv.factuurnummer,
        "factuurdatum": iso(inv.factuurdatum),
        "vervaldatum": iso(inv.vervaldatum),
        "klantnaam": inv.klantnaam,
        "regels": [
            {
                "datum": iso(ln.datum),
                "aantal_uren": float(ln.aantal_uren) if ln.aantal_uren else None,
                "tarief": float(ln.tarief) if ln.tarief else None,
                "omschrijving": ln.omschrijving,
            }
            for ln in lines
        ],
        "totals": totals,
    }


# ----------------------------
# AI Prompts - Core (existing)
# ----------------------------
def build_rewrite_prompt(inv: InvoiceHeader, lines: List[InvoiceLine]) -> str:
    items = []
    for ln in lines:
        dt = iso(ln.datum) or ""
        hrs = f"{ln.aantal_uren}".replace(".", ",") if ln.aantal_uren is not None else ""
        items.append(f"- {dt}: {hrs} uur — {ln.omschrijving}")
    raw = "\n".join(items)

    return f"""
Je schrijft een nette, zakelijke factuur-omschrijving in het Nederlands voor een ZZP IT/consultancy factuur.
Maak een korte kopregel en daarna maximaal 3 bullets. Houd het professioneel en concreet.

Klant: {inv.klantnaam}
Factuurnummer: {inv.factuurnummer}

Ruwe input (uit Excel):
{raw}

Geef ALLEEN de tekst terug (geen code fences).
""".strip()


def build_audit_prompt(inv: InvoiceHeader, lines: List[InvoiceLine], totals: Dict[str, Any], omschrijving: str) -> str:
    payload = {
        "factuurnummer": inv.factuurnummer,
        "factuurdatum": iso(inv.factuurdatum),
        "vervaldatum": iso(inv.vervaldatum),
        "debiteurnummer": inv.debiteurnummer,
        "klantnaam": inv.klantnaam,
        "klant_adres": inv.klant_adres,
        "klant_postcode_plaats": inv.klant_postcode_plaats,
        "regels": [
            {
                "datum": iso(ln.datum),
                "aantal_uren": float(ln.aantal_uren) if ln.aantal_uren is not None else None,
                "tarief": float(ln.tarief) if ln.tarief is not None else None,
                "btw_pct": float(ln.btw_pct) if ln.btw_pct is not None else None,
                "omschrijving": ln.omschrijving,
            }
            for ln in lines
        ],
        "totals": totals,
        "omschrijving": omschrijving,
    }

    return f"""
Je bent een strenge auditor voor Nederlandse ZZP facturen.
Controleer op ontbrekende velden, inconsistenties en onduidelijke omschrijving.
Geef ALLEEN JSON terug met exact dit schema:

{{
  "status": "OK" | "WARNING" | "ERROR",
  "warnings": [string],
  "errors": [string],
  "suggestions": [string]
}}

Data:
{json.dumps(payload, ensure_ascii=False, indent=2)}

Regels:
- ERROR bij ontbrekende klantnaam, factuurdatum, of geen regels.
- WARNING bij btw_pct niet 0/9/21, of lege omschrijving, of uren/tarief 0.
- Suggestion: korte verbetering (max 3).

ALLEEN JSON (geen code fences).
""".strip()


# ----------------------------
# AI Prompts - NEW FEATURES
# ----------------------------

# 1. Smart line item categorization
def build_categorize_prompt(lines: List[InvoiceLine]) -> str:
    items = [
        {"index": i, "omschrijving": ln.omschrijving, "datum": iso(ln.datum)}
        for i, ln in enumerate(lines)
    ]
    return f"""
Categoriseer elke factuurregel in EXACT één van deze categorieën:
{json.dumps(CATEGORIES)}

Input regels:
{json.dumps(items, ensure_ascii=False, indent=2)}

Geef ALLEEN een JSON array terug met voor elke regel een object:
{{"index": number, "category": string, "confidence": number (0-100)}}

Regels:
- "development": programmeren, bouwen, implementeren, coderen
- "consulting": advies, strategie, analyse, planning
- "testing": testen, QA, debugging, review
- "meeting": vergadering, overleg, call, presentatie
- "support": onderhoud, bugfix, helpdesk, troubleshooting
- "other": alles wat niet past

ALLEEN JSON array (geen code fences).
""".strip()


# 2. Client communication drafts
def build_email_prompt(inv: InvoiceHeader, totals: Dict[str, Any], omschrijving: str) -> str:
    return f"""
Schrijf een korte, vriendelijke e-mail in het Nederlands om deze factuur te versturen.

Factuurgegevens:
- Klant: {inv.klantnaam}
- Factuurnummer: {inv.factuurnummer}
- Bedrag: {totals['totaal_incl']}
- Vervaldatum: {iso(inv.vervaldatum) or "30 dagen"}

Werkzaamheden (samenvatting):
{omschrijving}

Regels:
- Professioneel maar niet stijf
- Maximaal 5 zinnen
- Begin met "Beste" of voornaam als die uit klantnaam te halen is
- Geen "Geachte heer/mevrouw"
- Eindig met vriendelijke groet
- Vermeld dat factuur als bijlage is toegevoegd

Geef ALLEEN de e-mailtekst terug.
""".strip()


# 3. Anomaly detection
def build_anomaly_prompt(
    inv: InvoiceHeader,
    lines: List[InvoiceLine],
    totals: Dict[str, Any],
    history: List[Dict[str, Any]]
) -> str:
    current = invoice_to_dict(inv, lines, totals)
    
    return f"""
Vergelijk deze factuur met de historische facturen van dezelfde klant.
Signaleer afwijkingen die aandacht verdienen.

HUIDIGE FACTUUR:
{json.dumps(current, ensure_ascii=False, indent=2)}

HISTORISCHE FACTUREN (zelfde klant):
{json.dumps(history, ensure_ascii=False, indent=2)}

Check op:
- Uurtarief significant hoger/lager dan normaal (>20% afwijking)
- Aantal uren veel meer/minder dan gemiddeld
- Nieuwe type werkzaamheden die niet eerder voorkwamen
- Ongebruikelijke factuurdatum (weekend, feestdag)
- Grote sprong in factuurbedrag

Geef ALLEEN JSON terug:
{{
  "has_anomalies": boolean,
  "anomalies": [
    {{"type": string, "description": string, "severity": "low"|"medium"|"high"}}
  ],
  "avg_previous_total": number|null,
  "avg_previous_hours": number|null,
  "typical_rate": number|null
}}

Als er geen historie is, geef has_anomalies: false en lege arrays.
ALLEEN JSON (geen code fences).
""".strip()


# 4. Payment reminder generation
def build_reminder_prompt(inv: InvoiceHeader, totals: Dict[str, Any], days_overdue: int, reminder_count: int) -> str:
    if reminder_count == 1:
        tone = "vriendelijk en begripvol"
        urgency = "licht"
    elif reminder_count == 2:
        tone = "zakelijk maar beleefd"
        urgency = "duidelijk"
    else:
        tone = "formeel en beslist"
        urgency = "urgent"

    return f"""
Schrijf een betalingsherinnering ({tone}) voor een openstaande factuur.

Factuurgegevens:
- Klant: {inv.klantnaam}
- Factuurnummer: {inv.factuurnummer}
- Bedrag: {totals['totaal_incl']}
- Originele vervaldatum: {iso(inv.vervaldatum)}
- Dagen over termijn: {days_overdue}
- Dit is herinnering nummer: {reminder_count}

Toon: {tone}
Urgentie: {urgency}

Regels:
- Maximaal 6 zinnen
- Geen dreigementen, wel duidelijk
- Herinnering 1: "mogelijk over het hoofd gezien"
- Herinnering 2: "verzoek om spoedige betaling"
- Herinnering 3+: "dringend verzoek, laatste herinnering voor verdere stappen"
- Vermeld altijd factuurnummer en bedrag
- Bied aan om contact op te nemen bij vragen

Geef ALLEEN de herinneringstekst terug.
""".strip()


# 5. Year-end summary
def build_yearly_summary_prompt(client_name: str, year: int, invoices_data: List[Dict[str, Any]]) -> str:
    return f"""
Genereer een professioneel jaaroverzicht voor een klant.

Klant: {client_name}
Jaar: {year}

Factuurdata:
{json.dumps(invoices_data, ensure_ascii=False, indent=2)}

Schrijf een overzicht met:
1. Totaal gefactureerd bedrag (excl. en incl. BTW)
2. Totaal aantal uren
3. Verdeling per kwartaal (in lopende tekst, geen tabel)
4. Top 3 type werkzaamheden (als te bepalen uit omschrijvingen)
5. Gemiddeld uurtarief
6. Korte samenvatting/bedankje (2-3 zinnen)

Format: nette markdown, maar geen tabellen.
Houd het professioneel en overzichtelijk.

Geef ALLEEN de markdown terug.
""".strip()


# ----------------------------
# Feature: Categorization
# ----------------------------
def run_categorize(client: OpenAI, model: str, inv: InvoiceHeader, lines: List[InvoiceLine]) -> Dict[str, Any]:
    if not lines:
        return {"categories": [], "summary": {}}
    
    prompt = build_categorize_prompt(lines)
    result = call_llm_json(client, model, prompt)
    
    # Build summary
    summary: Dict[str, float] = defaultdict(float)
    for item in result:
        cat = item.get("category", "other")
        idx = item.get("index", 0)
        if 0 <= idx < len(lines) and lines[idx].aantal_uren:
            summary[cat] += float(lines[idx].aantal_uren)
    
    return {
        "categories": result,
        "summary": dict(summary),
    }


# ----------------------------
# Feature: Email draft
# ----------------------------
def run_email(client: OpenAI, model: str, inv: InvoiceHeader, totals: Dict[str, Any], omschrijving: str) -> str:
    prompt = build_email_prompt(inv, totals, omschrijving)
    return call_llm_text(client, model, prompt)


# ----------------------------
# Feature: Anomaly detection
# ----------------------------
def run_anomaly(
    client: OpenAI,
    model: str,
    inv: InvoiceHeader,
    lines: List[InvoiceLine],
    totals: Dict[str, Any],
    all_invoices: List[InvoiceHeader],
    lines_by: Dict[str, List[InvoiceLine]]
) -> Dict[str, Any]:
    # Build history for same client
    history = []
    for other in all_invoices:
        if other.klantnaam == inv.klantnaam and other.factuurnummer != inv.factuurnummer:
            other_lines = lines_by.get(other.factuurnummer, [])
            other_totals = compute_totals(other_lines)
            history.append(invoice_to_dict(other, other_lines, other_totals))
    
    # Sort by date descending, limit to last 10
    history.sort(key=lambda x: x.get("factuurdatum") or "", reverse=True)
    history = history[:10]
    
    prompt = build_anomaly_prompt(inv, lines, totals, history)
    return call_llm_json(client, model, prompt)


# ----------------------------
# Feature: Payment reminder
# ----------------------------
def run_reminder(
    client: OpenAI,
    model: str,
    inv: InvoiceHeader,
    totals: Dict[str, Any],
    reminder_count: int = 1
) -> Optional[Dict[str, Any]]:
    if not inv.vervaldatum:
        return None
    
    today = date.today()
    if inv.vervaldatum >= today:
        return None  # Not overdue
    
    days_overdue = (today - inv.vervaldatum).days
    
    prompt = build_reminder_prompt(inv, totals, days_overdue, reminder_count)
    text = call_llm_text(client, model, prompt)
    
    return {
        "days_overdue": days_overdue,
        "reminder_number": reminder_count,
        "text": text,
    }


# ----------------------------
# Feature: Yearly summary
# ----------------------------
def run_yearly_summary(
    client: OpenAI,
    model: str,
    client_name: str,
    year: int,
    invoices: List[InvoiceHeader],
    lines_by: Dict[str, List[InvoiceLine]]
) -> str:
    # Filter invoices for this client and year
    filtered = [
        inv for inv in invoices
        if inv.klantnaam == client_name
        and inv.factuurdatum
        and inv.factuurdatum.year == year
    ]
    
    if not filtered:
        return f"Geen facturen gevonden voor {client_name} in {year}."
    
    # Build data for prompt
    invoices_data = []
    for inv in sorted(filtered, key=lambda x: x.factuurdatum or date.min):
        inv_lines = lines_by.get(inv.factuurnummer, [])
        totals = compute_totals(inv_lines)
        invoices_data.append({
            "factuurnummer": inv.factuurnummer,
            "factuurdatum": iso(inv.factuurdatum),
            "maand": inv.factuurdatum.month if inv.factuurdatum else None,
            "kwartaal": (inv.factuurdatum.month - 1) // 3 + 1 if inv.factuurdatum else None,
            "uren": totals["totaal_uren"],
            "bedrag_excl": totals["totaal_excl_float"],
            "bedrag_incl": totals["totaal_incl_float"],
            "omschrijvingen": [ln.omschrijving for ln in inv_lines],
        })
    
    prompt = build_yearly_summary_prompt(client_name, year, invoices_data)
    return call_llm_text(client, model, prompt)


# ----------------------------
# Main
# ----------------------------
def main() -> None:
    load_dotenv(ROOT / ".env")
    api_key = os.getenv("OPENAI_API_KEY")
    model = os.getenv("OPENAI_MODEL", "gpt-4o-mini")

    if not api_key:
        raise SystemExit("❌ OPENAI_API_KEY ontbreekt in .env")

    parser = argparse.ArgumentParser(
        description="invoice-ai v2: rewrite, audit, categorize, email, anomaly, reminder, summary",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Voorbeelden:
  # Standaard verwerking (rewrite + audit)
  python invoice_ai.py --excel facturen.xlsx --factuur BS-2025-001

  # Alle features voor één factuur
  python invoice_ai.py --excel facturen.xlsx --factuur BS-2025-001 --all-features

  # E-mail draft genereren
  python invoice_ai.py --excel facturen.xlsx --factuur BS-2025-001 --email

  # Betalingsherinnering (2e herinnering)
  python invoice_ai.py --excel facturen.xlsx --factuur BS-2025-001 --reminder --reminder-count 2

  # Jaaroverzicht voor klant
  python invoice_ai.py --excel facturen.xlsx --yearly-summary "Klant BV" --year 2024
        """
    )
    
    # Required
    parser.add_argument("--excel", required=True, help="Pad naar facturen.xlsx")
    
    # Invoice selection
    parser.add_argument("--factuur", default=None, help="Specifiek factuurnummer")
    
    # Feature flags
    parser.add_argument("--all-features", action="store_true", help="Run alle AI features")
    parser.add_argument("--categorize", action="store_true", help="Categoriseer regelomschrijvingen")
    parser.add_argument("--email", action="store_true", help="Genereer e-mail draft")
    parser.add_argument("--anomaly", action="store_true", help="Check op afwijkingen t.o.v. historie")
    parser.add_argument("--reminder", action="store_true", help="Genereer betalingsherinnering (als verlopen)")
    parser.add_argument("--reminder-count", type=int, default=1, help="Herinneringsnummer (1, 2, 3+)")
    
    # Yearly summary (separate mode)
    parser.add_argument("--yearly-summary", metavar="KLANTNAAM", help="Genereer jaaroverzicht voor klant")
    parser.add_argument("--year", type=int, default=None, help="Jaar voor jaaroverzicht (default: huidig jaar)")
    
    # Skip options
    parser.add_argument("--skip-rewrite", action="store_true", help="Skip rewrite stap")
    parser.add_argument("--skip-audit", action="store_true", help="Skip audit stap")
    parser.add_argument("--force", action="store_true", help="Herverwerk ook als output al bestaat")

    args = parser.parse_args()

    excel_path = Path(args.excel).resolve()
    if not excel_path.exists():
        raise SystemExit(f"❌ Excel niet gevonden: {excel_path}")

    invoices, lines = load_excel(excel_path)
    lines_by = group_lines(lines)
    client = OpenAI(api_key=api_key)

    # --- YEARLY SUMMARY MODE ---
    if args.yearly_summary:
        year = args.year or date.today().year
        print(f"📊 Jaaroverzicht voor {args.yearly_summary} ({year})...")
        
        summary = run_yearly_summary(client, model, args.yearly_summary, year, invoices, lines_by)
        
        safe_name = "".join(c if c.isalnum() else "_" for c in args.yearly_summary)
        out_path = OUT_DIR / f"yearly_{safe_name}_{year}.md"
        out_path.write_text(summary, encoding="utf-8")
        
        print(f"✅ Jaaroverzicht: {out_path}")
        print("\n" + summary)
        return

    # --- INVOICE PROCESSING MODE ---
    if args.all_features:
        args.categorize = True
        args.email = True
        args.anomaly = True
        args.reminder = True

    # Select invoices
    selected = invoices
    if args.factuur:
        selected = [inv for inv in invoices if inv.factuurnummer == args.factuur]
        if not selected:
            raise SystemExit(f"❌ Factuurnummer niet gevonden: {args.factuur}")

    for inv in selected:
        print(f"\n{'='*50}")
        print(f"📄 {inv.factuurnummer} - {inv.klantnaam}")
        print(f"{'='*50}")
        
        inv_lines = lines_by.get(inv.factuurnummer, [])
        if not inv_lines:
            print(f"⚠️  Geen regels gevonden")
        
        totals = compute_totals(inv_lines)
        omschrijving = ""

        # --- REWRITE ---
        if not args.skip_rewrite:
            rewrite_path = OUT_DIR / f"{inv.factuurnummer}.rewrite.txt"
            if not args.force and rewrite_path.exists():
                print(f"⏭️  Rewrite: al aanwezig")
                omschrijving = rewrite_path.read_text(encoding="utf-8")
            else:
                print(f"✍️  Rewrite...")
                omschrijving = call_llm_text(client, model, build_rewrite_prompt(inv, inv_lines))
                rewrite_path.write_text(omschrijving, encoding="utf-8")
                print(f"   → {rewrite_path}")

        # --- AUDIT ---
        if not args.skip_audit:
            audit_path = OUT_DIR / f"{inv.factuurnummer}.audit.json"
            if not args.force and audit_path.exists():
                print(f"⏭️  Audit: al aanwezig")
                audit = json.loads(audit_path.read_text(encoding="utf-8"))
            else:
                print(f"🔍 Audit...")
                audit = call_llm_json(client, model, build_audit_prompt(inv, inv_lines, totals, omschrijving))
                audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
                print(f"   → {audit_path}")
            
            status = audit.get("status", "?")
            print(f"   Status: {status}")
            for w in audit.get("warnings", []):
                print(f"   ⚠️  {w}")
            for e in audit.get("errors", []):
                print(f"   ❌ {e}")

        # --- CATEGORIZE ---
        if args.categorize:
            cat_path = OUT_DIR / f"{inv.factuurnummer}.categories.json"
            if not args.force and cat_path.exists():
                print(f"⏭️  Categorize: al aanwezig")
            else:
                print(f"🏷️  Categorize...")
                cat_result = run_categorize(client, model, inv, inv_lines)
                cat_path.write_text(json.dumps(cat_result, ensure_ascii=False, indent=2), encoding="utf-8")
                print(f"   → {cat_path}")
                if cat_result.get("summary"):
                    print(f"   Uren per categorie: {cat_result['summary']}")

        # --- EMAIL ---
        if args.email:
            email_path = OUT_DIR / f"{inv.factuurnummer}.email.txt"
            if not args.force and email_path.exists():
                print(f"⏭️  Email: al aanwezig")
            else:
                print(f"📧 Email draft...")
                if not omschrijving:
                    omschrijving = call_llm_text(client, model, build_rewrite_prompt(inv, inv_lines))
                email_text = run_email(client, model, inv, totals, omschrijving)
                email_path.write_text(email_text, encoding="utf-8")
                print(f"   → {email_path}")

        # --- ANOMALY ---
        if args.anomaly:
            anom_path = OUT_DIR / f"{inv.factuurnummer}.anomaly.json"
            if not args.force and anom_path.exists():
                print(f"⏭️  Anomaly: al aanwezig")
            else:
                print(f"🔎 Anomaly check...")
                anom_result = run_anomaly(client, model, inv, inv_lines, totals, invoices, lines_by)
                anom_path.write_text(json.dumps(anom_result, ensure_ascii=False, indent=2), encoding="utf-8")
                print(f"   → {anom_path}")
                if anom_result.get("has_anomalies"):
                    for a in anom_result.get("anomalies", []):
                        sev = a.get("severity", "?")
                        desc = a.get("description", "")
                        icon = "🔴" if sev == "high" else "🟡" if sev == "medium" else "🟢"
                        print(f"   {icon} {desc}")
                else:
                    print(f"   ✅ Geen afwijkingen gedetecteerd")

        # --- REMINDER ---
        if args.reminder:
            rem_path = OUT_DIR / f"{inv.factuurnummer}.reminder_{args.reminder_count}.txt"
            if not args.force and rem_path.exists():
                print(f"⏭️  Reminder: al aanwezig")
            else:
                print(f"⏰ Reminder check...")
                rem_result = run_reminder(client, model, inv, totals, args.reminder_count)
                if rem_result:
                    rem_path.write_text(rem_result["text"], encoding="utf-8")
                    print(f"   → {rem_path}")
                    print(f"   Dagen over termijn: {rem_result['days_overdue']}")
                else:
                    print(f"   ℹ️  Niet verlopen of geen vervaldatum")

    print(f"\n{'='*50}")
    print(f"✅ Klaar! Output in: {OUT_DIR}")


if __name__ == "__main__":
    main()
